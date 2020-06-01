#import modules
import matplotlib.pyplot as plt
import numpy as np
import cv2
import utils
import openpyxl as pyx
import collections



#### Loop to input files ####

#create list of id numbers of all pictures
ids = ["03", "04", "05", "06", "07", "14", "15", "17", "18", "37", "38", "55"]

#define dict to collect insertion depth for every electroc
angles=collections.defaultdict(list)

#loop over all id numbers
for i in ids:
    
    #set pre and post images
    pre = plt.imread('DATA/ID' + i + '/ID' + i + 'pre.png')
    post = plt.imread('DATA/ID' + i + '/ID' + i + 'post.png')
    
    #crop out text of images and normalize pre and post images to homogenize color scale
    pre_crop = utils.crop_image(pre)
    post_crop = utils.crop_image(post)
    post_norm = utils.normalize(post_crop)
    
    #find spiral center
    thresh_img = utils.find_bright_points(post_norm)
    comp_img = utils.find_components(thresh_img)
    x_center, y_center = utils.find_center(comp_img)
    center = np.array([x_center, y_center])
    plt.plot(x_center, y_center, 'r.', markersize=14)

    thresh_img_spiral = utils.select_spiral(thresh_img, center)
    elec, cnts = utils.find_electrodes(post_norm, thresh_img_spiral, center)

    #ennumerate electrodes 
    enum_electrodes = utils.enumerate_electrodes(cnts, center, thresh_img)

    #plot center and electrode number to post-operative picture
    enum_electrodes.reverse()
    for elec in range(len(enum_electrodes)):
        cv2.putText(post_norm, str(elec+1), (int(enum_electrodes[elec][0]), int(enum_electrodes[elec][1])), cv2.FONT_HERSHEY_SIMPLEX, 2, (0,0,225), 3)             
    plt.imshow(post_norm)
    plt.show()
    
    #filter out additional points detected not belonging to 12 piece set of electrodes 
    last=len(enum_electrodes)
    if last>12:
        enum_electrodes=enum_electrodes[:12]
        last=len(enum_electrodes)
    
    #find individual angles between points
    ref = utils.find_insertion_angle(center, enum_electrodes[-1], enum_electrodes[-1])
    print('For ID' + str(i) + ' angles are:')
    print(last, 'New angle: {:.2f}'.format(abs(ref)), 'Total insertion angle: {:.2f}'.format(ref))
    #add angles to histogram dict
    angles[last].append(ref)
    
    #add up angles for insertion depth
    output={last:ref}
    for j in range(len(enum_electrodes)-1):
        angle = (utils.find_insertion_angle(center, enum_electrodes[-j], enum_electrodes[-j+1]))
        ref += abs(angle)
        number = last-j-1
        print(number, 'New angle: {:.2f}'.format(abs(angle)), 'Total insertion angle: {:.2f}'.format(ref))
        output[number]=ref
        #add angles to histogram dict
        angles[number].append(ref)

 


    #set the output coordinates and angles for every sample to print to outputsheet
  
    # Call a Workbook() function of openpyxl to open ore nade result sheet
    wb = pyx.load_workbook(filename = 'results_IBMAJA.xlsx')
  
    # Get workbook active sheet   
    sheet = wb.active 
  
    #define starting cell for every ID
    ID = {"03":[6,2], "04":[6,7], "05":[6,12], "06":[6,17], "07":[6,22], "14":[6,27], "15":[26,2], "17":[26,7], "18":[26,12], "37":[26,17], "38":[26,22], "55":[26,27]}
  
      
    #get location of specific cells
    center_x_row=ID[i][0]
    center_x_col=ID[i][1]
    
    center_y_row=ID[i][0]
    center_y_col=ID[i][1]+1
    
    top_elec_x_row=ID[i][0]+4
    top_elec_x_col=ID[i][1]
    
    top_elec_y_row=ID[i][0]+4
    top_elec_y_col=ID[i][1]+1
    
    top_angle_row=ID[i][0]+4
    top_angle_col=ID[i][1]+2
    
    #input values to cells
    sheet.cell(row=center_x_row, column=center_x_col).value=center[0]
    sheet.cell(row=center_y_row, column=center_y_col).value=center[1]

    
    for k in range(len(enum_electrodes)):
        #sheet[electrodes].value=k
        sheet.cell(row=top_elec_x_row+k, column=top_elec_x_col).value=enum_electrodes[k][0]
        sheet.cell(row=top_elec_y_row+k, column=top_elec_y_col).value=enum_electrodes[k][1]
        sheet.cell(row=top_angle_row+k, column=top_angle_col).value=output[k+1]

    #save in file
    wb.save("results_IBMAJA.xlsx")

means={}
for o in range(1,13):
    mean=(np.mean(angles[o]))
    means[o]=mean

print(means)

tick=(1,2,3,4,5,6,7,8,9,10,11,12)   
plt.title('Barplot of the mean insertion angle of each electrode')
plt.xlabel('Electrode number')
plt.ylabel('Insertion angle') 
plt.bar(list(means.keys()), means.values(), color='b', tick_label=tick)
plt.show()
